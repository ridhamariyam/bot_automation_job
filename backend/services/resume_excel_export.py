"""
Resume Excel exporter — produces a structured .xlsx with one sheet per section.
"""
import io
import json
import logging
from typing import Any

logger = logging.getLogger(__name__)


def export_resume_excel(resume_data: dict) -> bytes:
    """Return Excel file bytes for the given resume data dict."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    wb = Workbook()

    HEADER_FILL  = PatternFill("solid", fgColor="1E40AF")   # blue-800
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    BOLD         = Font(bold=True)
    CENTER       = Alignment(horizontal="center", vertical="center")
    THIN         = Side(style="thin", color="CBD5E1")
    BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _style_header_row(ws, columns: list[str]):
        ws.append(columns)
        for cell in ws[ws.max_row]:
            cell.fill    = HEADER_FILL
            cell.font    = HEADER_FONT
            cell.alignment = CENTER
            cell.border  = BORDER

    def _auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    def _bullets(raw) -> str:
        if not raw:
            return ""
        if isinstance(raw, list):
            return "\n".join(f"• {b}" for b in raw)
        try:
            parsed = json.loads(raw)
            return "\n".join(f"• {b}" for b in parsed)
        except Exception:
            return str(raw)

    # ── Sheet 1: Personal Info ─────────────────────────────────────────────────
    ws_info = wb.active
    ws_info.title = "Personal Info"
    _style_header_row(ws_info, ["Field", "Value"])
    for field, value in [
        ("Full Name",            resume_data.get("full_name", "")),
        ("Email",                resume_data.get("email", "")),
        ("Phone",                resume_data.get("phone", "")),
        ("Location",             resume_data.get("location", "")),
        ("LinkedIn",             resume_data.get("linkedin_url", "")),
        ("GitHub",               resume_data.get("github_url", "")),
        ("Website",              resume_data.get("website_url", "")),
        ("Professional Summary", resume_data.get("professional_summary", "")),
    ]:
        row = ws_info.append([field, value]) or ws_info[ws_info.max_row]
        for cell in ws_info[ws_info.max_row]:
            cell.border = BORDER
    _auto_width(ws_info)

    # ── Sheet 2: Experience ────────────────────────────────────────────────────
    ws_exp = wb.create_sheet("Experience")
    _style_header_row(ws_exp, ["Title", "Company", "Location", "Start", "End / Current", "Bullets"])
    for exp in resume_data.get("experiences", []):
        ws_exp.append([
            exp.get("title", ""),
            exp.get("company", ""),
            exp.get("location", ""),
            exp.get("start_date", ""),
            "Present" if exp.get("current") else exp.get("end_date", ""),
            _bullets(exp.get("bullets")),
        ])
        for cell in ws_exp[ws_exp.max_row]:
            cell.border    = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws_exp.row_dimensions[1].height = 20
    _auto_width(ws_exp)

    # ── Sheet 3: Skills ────────────────────────────────────────────────────────
    ws_sk = wb.create_sheet("Skills")
    _style_header_row(ws_sk, ["Category", "Skill", "Proficiency"])
    for sk in resume_data.get("skills", []):
        ws_sk.append([sk.get("category", ""), sk.get("skill", ""), sk.get("proficiency", "")])
        for cell in ws_sk[ws_sk.max_row]:
            cell.border = BORDER
    _auto_width(ws_sk)

    # ── Sheet 4: Projects ─────────────────────────────────────────────────────
    ws_proj = wb.create_sheet("Projects")
    _style_header_row(ws_proj, ["Name", "Tech Stack", "URL", "Description / Bullets"])
    for proj in resume_data.get("projects", []):
        bullets_text = _bullets(proj.get("bullets")) or proj.get("description", "")
        ws_proj.append([
            proj.get("name", ""),
            proj.get("tech_stack", ""),
            proj.get("url", ""),
            bullets_text,
        ])
        for cell in ws_proj[ws_proj.max_row]:
            cell.border    = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _auto_width(ws_proj)

    # ── Sheet 5: Education ────────────────────────────────────────────────────
    ws_edu = wb.create_sheet("Education")
    _style_header_row(ws_edu, ["Institution", "Degree", "Field", "Start", "End", "GPA", "Achievements"])
    for edu in resume_data.get("educations", []):
        ws_edu.append([
            edu.get("institution", ""),
            edu.get("degree", ""),
            edu.get("field", ""),
            edu.get("start_year", ""),
            edu.get("end_year", ""),
            edu.get("gpa", ""),
            edu.get("achievements", ""),
        ])
        for cell in ws_edu[ws_edu.max_row]:
            cell.border = BORDER
    _auto_width(ws_edu)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
